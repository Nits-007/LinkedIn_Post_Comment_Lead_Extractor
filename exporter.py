import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


COLUMNS = [
    ("extracted_at", 22),
    ("post_url", 50),
    ("commenter_name", 25),
    ("commenter_profile_url", 45),
    ("comment_text", 60),
    ("email", 30),
]


def export_to_excel(leads: list[dict], output_path: str) -> str:

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Leads"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0B5394", end_color="0B5394", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_width


    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, lead in enumerate(leads, start=2):
        for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
            value = lead.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    logger.info(f"Exported {len(leads)} leads to: {abs_path}")
    return abs_path
